from flask import Flask, render_template, request, redirect, url_for, send_file
from random import choice
import os
from openpyxl import load_workbook
import io

app = Flask(__name__)
app.secret_key = 'your-secret-key-123'
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

participants = []

def process_names_input(names_input):
    """处理手动输入的名字"""
    return [name.strip() for name in names_input.split() if name.strip()]

def process_uploaded_file(file):
    """处理上传的文件"""
    names = []
    if not file or not file.filename:
        return names

    filename = file.filename.lower()
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)

    try:
        file.save(filepath)

        if filename.endswith('.txt'):
            with open(filepath, 'r', encoding='utf-8') as f:
                names = [line.strip() for line in f if line.strip()]

        elif filename.endswith('.xlsx'):
            wb = load_workbook(filepath)
            sheet = wb.active

            # 获取第一列的所有非空单元格
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:  # 只取第一列的数据
                    name = str(row[0]).strip()
                    if name:
                        names.append(name)

        return names

    except Exception as e:
        print(f"文件处理错误: {e}")
        return []

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

@app.route('/')
def home():
    names_list = "、".join(participants) if participants else None
    return render_template('index_2.html',
                         names_list=names_list,
                         count=len(participants))

@app.route('/set_names', methods=['POST'])
def set_names():
    global participants

    # 处理手动输入
    names_input = request.form.get('names', '').strip()
    new_names = process_names_input(names_input)

    # 处理文件上传
    uploaded_file = request.files.get('file')
    if uploaded_file and uploaded_file.filename:
        new_names += process_uploaded_file(uploaded_file)

    if new_names:
        participants = new_names

    return redirect(url_for('home'))

@app.route('/clear')
def clear_list():
    global participants
    participants = []
    return redirect(url_for('home'))

@app.route('/shuiji')
def draw_winner():
    if not participants:
        return redirect(url_for('home'))
    winner = choice(participants)
    names_list = "、".join(participants)
    return render_template('index.html',
                         names_list=names_list,
                         count=len(participants),
                         result=winner)

@app.route('/export')
def export_names():
    """导出名单为TXT文件"""
    if not participants:
        return redirect(url_for('home'))

    # 创建内存中的文件
    file_data = io.StringIO()
    file_data.write("\n".join(participants))
    file_data.seek(0)

    return send_file(
        io.BytesIO(file_data.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name='participants.txt'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)